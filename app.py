from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Dict, List, Optional

import pandas as pd
import requests
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# SharePoint auto-load URL
# Converted from the share link to a direct-download link by appending
# &download=1. If the file is behind Atlan SSO this will fall back to
# manual upload automatically.
# ---------------------------------------------------------------------------
COMPETITOR_SHAREPOINT_URL = (
    "https://atlanstormwater.sharepoint.com/:x:/s/CIA/"
    "IQASFO4gDYM5T4hmHT4y1q3HARdKM1KOlS9IlW3sY_43GCM?e=zd7jF1&download=1"
)

st.set_page_config(page_title="Atlan Pricing Engine", page_icon="💧", layout="wide")

# ---------------------------------------------------------------------------
# Static data
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class Region:
    name: str
    market_factor: float
    freight_factor: float
    notes: str


@dataclass(frozen=True)
class Fleet:
    name: str
    litres_per_100km: float
    maintenance_per_km: float


# State -> column name mapping for NetSuite price list
STATE_TO_NETSUITE_COL = {
    "VIC": "VIC",
    "NSW": "NSW / ACT",
    "QLD": "QLD",
    "WA": "WA",
    "SA": "SA",
    "TAS": "TAS",
    "NT": "NT",
}

# Competitor Intelligence: regions map to the combined State choice values
# used in the Power Apps form (e.g. "VIC/TAS", "QLD/NT", "NSW/ACT")
COMPETITOR_STATE_MAP = {
    "VIC": ["VIC/TAS", "VIC"],
    "TAS": ["VIC/TAS", "TAS"],
    "NSW": ["NSW/ACT", "NSW"],
    "QLD": ["QLD/NT", "QLD"],
    "WA": ["WA"],
    "SA": ["SA"],
    "NT": ["QLD/NT", "NT"],
}

REGIONS: Dict[str, Region] = {
    "VIC": Region("Victoria", 1.00, 1.00, "Balanced market with room for value-led pricing."),
    "QLD": Region("Queensland", 0.95, 1.05, "Competitive market with stronger price pressure."),
    "NSW": Region("New South Wales", 0.97, 1.10, "High-volume market with active peer competition."),
    "WA": Region("Western Australia", 1.05, 1.18, "Higher freight exposure and supply cost."),
    "SA": Region("South Australia", 1.02, 1.12, "Moderate pricing pressure with freight sensitivity."),
    "TAS": Region("Tasmania", 1.06, 1.30, "Freight-sensitive market with delivery complexity."),
    "NT": Region("Northern Territory", 0.94, 1.20, "Remote market with higher freight sensitivity."),
}

FLEET: Dict[str, Fleet] = {
    "Ute": Fleet("Ute", 12, 0.10),
    "Ute + trailer": Fleet("Ute + trailer", 16, 0.14),
    "6.5m truck": Fleet("6.5m truck", 20, 0.18),
    "6.5m truck + trailer": Fleet("6.5m truck + trailer", 25, 0.22),
    "8m truck": Fleet("8m truck", 30, 0.22),
    "8m truck + trailer": Fleet("8m truck + trailer", 35, 0.28),
}

ZONES = {
    "Metro": 30,
    "Outer Metro": 60,
    "Regional": 150,
    "Remote": 350,
    "TAS": 600,
}

DISCOUNT_OPTIONS = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
COST_FACTOR = 0.65  # default product cost as fraction of RRP/sell price, used when no manual override

# ---------------------------------------------------------------------------
# Built-in price list (ATF product range)
# ---------------------------------------------------------------------------
_BUILTIN_PRICE_DATA = [
    {"Internal ID": 8766,  "Name": "ATF1050-CAP",         "Type": "Assembly",        "Display Name": "AtlanFlow 1050mm SN8 Cap Fitting",              "Base Price": None,   "NSW / ACT": 9279,   "NT": 9279,   "QLD": 9279,   "SA": 9279,   "TAS": 9279,   "VIC": 9279,   "WA": 9279},
    {"Internal ID": 8175,  "Name": "ATF1050.8",            "Type": "Assembly",        "Display Name": "AtlanFlow DN1050 SN8",                          "Base Price": 2091.7, "NSW / ACT": 2099.6, "NT": 2099.6, "QLD": 2091.7, "SA": 2191,   "TAS": 2191,   "VIC": 2091.7, "WA": None},
    {"Internal ID": 10190, "Name": "ATF1050.8-K",          "Type": "Kit/Package",     "Display Name": "ATF1050.8-K",                                   "Base Price": 1500,   "NSW / ACT": 1600,   "NT": 1576,   "QLD": 1546,   "SA": 2748,   "TAS": 2134,   "VIC": 8787,   "WA": None},
    {"Internal ID": 9942,  "Name": "ATF1050.8-PERF",       "Type": "Assembly",        "Display Name": "DN1050 SN8 Atlan Flow Perforated",              "Base Price": 199.9,  "NSW / ACT": 157.8,  "NT": 157.8,  "QLD": 199.9,  "SA": 211,    "TAS": 211,    "VIC": 119.9,  "WA": None},
    {"Internal ID": 8244,  "Name": "ATF1050.GASKET",       "Type": "Inventory Item",  "Display Name": "DN1050 GASKET",                                 "Base Price": 150,    "NSW / ACT": 150,    "NT": 150,    "QLD": 150,    "SA": 150,    "TAS": 150,    "VIC": 150,    "WA": 150},
    {"Internal ID": 8713,  "Name": "ATF110-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 110mm SN8 45 Degree Fitting",         "Base Price": 309,    "NSW / ACT": 309,    "NT": 309,    "QLD": 309,    "SA": 309,    "TAS": 309,    "VIC": 309,    "WA": 309},
    {"Internal ID": 8714,  "Name": "ATF110-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 110mm SN8 90 Degree Fitting",         "Base Price": 379,    "NSW / ACT": 379,    "NT": 379,    "QLD": 379,    "SA": 379,    "TAS": 379,    "VIC": 379,    "WA": 379},
    {"Internal ID": 8092,  "Name": "ATF110.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN110 SN8",                           "Base Price": 40,     "NSW / ACT": 39,     "NT": 39,     "QLD": 40,     "SA": 40,     "TAS": 40,     "VIC": 40,     "WA": 40},
    {"Internal ID": 8484,  "Name": "ATF110.8-PERF",        "Type": "Assembly",        "Display Name": "DN110 SN8 Atlan Flow Perforated",               "Base Price": 40,     "NSW / ACT": 39,     "NT": 39,     "QLD": 40,     "SA": 40,     "TAS": 40,     "VIC": 40,     "WA": 40},
    {"Internal ID": 8362,  "Name": "ATF110.GASKET",        "Type": "Inventory Item",  "Display Name": "DN110 GASKET",                                  "Base Price": 15,     "NSW / ACT": 15,     "NT": 15,     "QLD": 15,     "SA": 15,     "TAS": 15,     "VIC": 15,     "WA": 15},
    {"Internal ID": 9909,  "Name": "ATF1200.8",            "Type": "Assembly",        "Display Name": "AtlanFlow 1200 SN8",                            "Base Price": None,   "NSW / ACT": None,   "NT": None,   "QLD": None,   "SA": None,   "TAS": None,   "VIC": None,   "WA": None},
    {"Internal ID": 8718,  "Name": "ATF160-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 160mm SN8 45 Degree Fitting",         "Base Price": 342,    "NSW / ACT": 342,    "NT": 342,    "QLD": 342,    "SA": 342,    "TAS": 342,    "VIC": 342,    "WA": 342},
    {"Internal ID": 8719,  "Name": "ATF160-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 160mm SN8 90 Degree Fitting",         "Base Price": 432,    "NSW / ACT": 432,    "NT": 432,    "QLD": 432,    "SA": 432,    "TAS": 432,    "VIC": 432,    "WA": 432},
    {"Internal ID": 8098,  "Name": "ATF160.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN160 SN8",                           "Base Price": 73,     "NSW / ACT": 45.8,   "NT": 45.8,   "QLD": 73,     "SA": 70,     "TAS": 70,     "VIC": 73,     "WA": 73},
    {"Internal ID": 8483,  "Name": "ATF160.8-PERF",        "Type": "Assembly",        "Display Name": "DN160 SN8 Atlan Flow Perforated",               "Base Price": 73,     "NSW / ACT": 45.8,   "NT": 45.8,   "QLD": 73,     "SA": 70,     "TAS": 70,     "VIC": 73,     "WA": 73},
    {"Internal ID": 8321,  "Name": "ATF160.GASKET",        "Type": "Inventory Item",  "Display Name": "DN160 GASKET",                                  "Base Price": 20,     "NSW / ACT": 20,     "NT": 20,     "QLD": 20,     "SA": 20,     "TAS": 20,     "VIC": 20,     "WA": 20},
    {"Internal ID": 8723,  "Name": "ATF225-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 225mm SN8 45 Degree Fitting",         "Base Price": 454,    "NSW / ACT": 454,    "NT": 454,    "QLD": 454,    "SA": 454,    "TAS": 454,    "VIC": 454,    "WA": 454},
    {"Internal ID": 8724,  "Name": "ATF225-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 225mm SN8 90 Degree Fitting",         "Base Price": 554,    "NSW / ACT": 554,    "NT": 554,    "QLD": 554,    "SA": 554,    "TAS": 554,    "VIC": 554,    "WA": 554},
    {"Internal ID": 7943,  "Name": "ATF225.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN225 SN8",                           "Base Price": 109.8,  "NSW / ACT": 99.5,   "NT": 99.5,   "QLD": 109.8,  "SA": 138,    "TAS": 138,    "VIC": 109.8,  "WA": 109.8},
    {"Internal ID": 9170,  "Name": "ATF225.8-PERF",        "Type": "Assembly",        "Display Name": "DN225 SN8 Atlan Flow Perforated",               "Base Price": 109.8,  "NSW / ACT": 99.5,   "NT": 99.5,   "QLD": 109.8,  "SA": 138,    "TAS": 138,    "VIC": 109.8,  "WA": 109.8},
    {"Internal ID": 8245,  "Name": "ATF225.GASKET",        "Type": "Inventory Item",  "Display Name": "DN225 GASKET",                                  "Base Price": 25,     "NSW / ACT": 25,     "NT": 25,     "QLD": 25,     "SA": 25,     "TAS": 25,     "VIC": 25,     "WA": 25},
    {"Internal ID": 8728,  "Name": "ATF300-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 300mm SN8 45 Degree Fitting",         "Base Price": 619,    "NSW / ACT": 619,    "NT": 619,    "QLD": 619,    "SA": 619,    "TAS": 619,    "VIC": 619,    "WA": 619},
    {"Internal ID": 8729,  "Name": "ATF300-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 300mm SN8 90 Degree Fitting",         "Base Price": 782,    "NSW / ACT": 782,    "NT": 782,    "QLD": 782,    "SA": 782,    "TAS": 782,    "VIC": 782,    "WA": 782},
    {"Internal ID": 7969,  "Name": "ATF300.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN300 SN8",                           "Base Price": 199.9,  "NSW / ACT": 157.8,  "NT": 157.8,  "QLD": 199.9,  "SA": 211,    "TAS": 211,    "VIC": 199.9,  "WA": 199.9},
    {"Internal ID": 8546,  "Name": "ATF300.8-PERF",        "Type": "Assembly",        "Display Name": "DN300 SN8 Atlan Flow Perforated",               "Base Price": 199.9,  "NSW / ACT": 157.8,  "NT": 157.8,  "QLD": 199.9,  "SA": 211,    "TAS": 211,    "VIC": 119.9,  "WA": 199.9},
    {"Internal ID": 7994,  "Name": "ATF300.GASKET",        "Type": "Inventory Item",  "Display Name": "DN300 GASKET",                                  "Base Price": 30,     "NSW / ACT": 30,     "NT": 30,     "QLD": 30,     "SA": 30,     "TAS": 30,     "VIC": 30,     "WA": 30},
    {"Internal ID": 8733,  "Name": "ATF375-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 375mm SN8 45 Degree Fitting",         "Base Price": 829,    "NSW / ACT": 829,    "NT": 829,    "QLD": 829,    "SA": 829,    "TAS": 829,    "VIC": 829,    "WA": 829},
    {"Internal ID": 8734,  "Name": "ATF375-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 375mm SN8 90 Degree Fitting",         "Base Price": 1049,   "NSW / ACT": 1049,   "NT": 1049,   "QLD": 1049,   "SA": 1049,   "TAS": 1049,   "VIC": 1049,   "WA": 1049},
    {"Internal ID": 8017,  "Name": "ATF375.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN375 SN8",                           "Base Price": 317.8,  "NSW / ACT": 231.6,  "NT": 231.6,  "QLD": 317.8,  "SA": 345,    "TAS": 345,    "VIC": 317.8,  "WA": 317.8},
    {"Internal ID": 9175,  "Name": "ATF375.8-PERF",        "Type": "Assembly",        "Display Name": "DN375 SN8 Atlan Flow Perforated",               "Base Price": 317.8,  "NSW / ACT": 231.6,  "NT": 231.6,  "QLD": 317.8,  "SA": 345,    "TAS": 345,    "VIC": 317.8,  "WA": 317.8},
    {"Internal ID": 8238,  "Name": "ATF375.GASKET",        "Type": "Inventory Item",  "Display Name": "DN375 GASKET",                                  "Base Price": 35,     "NSW / ACT": 35,     "NT": 35,     "QLD": 35,     "SA": 35,     "TAS": 35,     "VIC": 35,     "WA": 35},
    {"Internal ID": 8738,  "Name": "ATF450-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 450mm SN8 45 Degree Fitting",         "Base Price": 1151,   "NSW / ACT": 1151,   "NT": 1151,   "QLD": 1151,   "SA": 1151,   "TAS": 1151,   "VIC": 1151,   "WA": 1151},
    {"Internal ID": 8739,  "Name": "ATF450-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 450mm SN8 90 Degree Fitting",         "Base Price": 1401,   "NSW / ACT": 1401,   "NT": 1401,   "QLD": 1401,   "SA": 1401,   "TAS": 1401,   "VIC": 1401,   "WA": 1401},
    {"Internal ID": 8024,  "Name": "ATF450.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN450 SN8",                           "Base Price": 411.4,  "NSW / ACT": 377.1,  "NT": 377.1,  "QLD": 411.4,  "SA": 485,    "TAS": 485,    "VIC": 411.4,  "WA": 411.4},
    {"Internal ID": 9171,  "Name": "ATF450.8-PERF",        "Type": "Assembly",        "Display Name": "DN450 SN8 Atlan Flow Perforated",               "Base Price": 411.4,  "NSW / ACT": 377.1,  "NT": 377.1,  "QLD": 411.4,  "SA": 485,    "TAS": 485,    "VIC": 411.4,  "WA": 411.4},
    {"Internal ID": 8239,  "Name": "ATF450.GASKET",        "Type": "Inventory Item",  "Display Name": "DN450 GASKET",                                  "Base Price": 40,     "NSW / ACT": 40,     "NT": 40,     "QLD": 40,     "SA": 40,     "TAS": 40,     "VIC": 40,     "WA": 40},
    {"Internal ID": 8743,  "Name": "ATF525-45",            "Type": "Assembly",        "Display Name": "AtlanFlow 525mm SN8 45 Degree Fitting",         "Base Price": 1478,   "NSW / ACT": 1478,   "NT": 1478,   "QLD": 1478,   "SA": 1478,   "TAS": 1478,   "VIC": 1478,   "WA": 1478},
    {"Internal ID": 8744,  "Name": "ATF525-90",            "Type": "Assembly",        "Display Name": "AtlanFlow 525mm SN8 90 Degree Fitting",         "Base Price": 1750,   "NSW / ACT": 1750,   "NT": 1750,   "QLD": 1750,   "SA": 1750,   "TAS": 1750,   "VIC": 1750,   "WA": 1750},
    {"Internal ID": 8103,  "Name": "ATF525.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN525 SN8",                           "Base Price": 527,    "NSW / ACT": 477.1,  "NT": 477.1,  "QLD": 527,    "SA": 648,    "TAS": 648,    "VIC": 527,    "WA": 527},
    {"Internal ID": 9172,  "Name": "ATF525.8-PERF",        "Type": "Assembly",        "Display Name": "DN525 SN8 Atlan Flow Perforated",               "Base Price": 527,    "NSW / ACT": 477.1,  "NT": 477.1,  "QLD": 527,    "SA": 648,    "TAS": 648,    "VIC": 527,    "WA": 527},
    {"Internal ID": 8240,  "Name": "ATF525.GASKET",        "Type": "Inventory Item",  "Display Name": "DN525 GASKET",                                  "Base Price": 45,     "NSW / ACT": 45,     "NT": 45,     "QLD": 45,     "SA": 45,     "TAS": 45,     "VIC": 45,     "WA": 45},
    {"Internal ID": 8074,  "Name": "ATF600.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN600 SN8",                           "Base Price": 773.1,  "NSW / ACT": 692.2,  "NT": 692.2,  "QLD": 773.1,  "SA": 820,    "TAS": 820,    "VIC": 773.1,  "WA": 773.1},
    {"Internal ID": 9173,  "Name": "ATF600.8-PERF",        "Type": "Assembly",        "Display Name": "DN600 SN8 Atlan Flow Perforated",               "Base Price": 773.1,  "NSW / ACT": 692.2,  "NT": 692.2,  "QLD": 773.1,  "SA": 820,    "TAS": 820,    "VIC": 773.1,  "WA": 773.1},
    {"Internal ID": 8241,  "Name": "ATF600.GASKET",        "Type": "Inventory Item",  "Display Name": "DN600 GASKET",                                  "Base Price": 50,     "NSW / ACT": 50,     "NT": 50,     "QLD": 50,     "SA": 50,     "TAS": 50,     "VIC": 50,     "WA": 50},
    {"Internal ID": 8193,  "Name": "ATF750.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN750 SN8",                           "Base Price": 1012.3, "NSW / ACT": 1098.1, "NT": 1098.1, "QLD": 1012.3, "SA": 1182,   "TAS": 1182,   "VIC": 1012.3, "WA": 1012.3},
    {"Internal ID": 9174,  "Name": "ATF750.8-PERF",        "Type": "Assembly",        "Display Name": "DN750 SN8 Atlan Flow Perforated",               "Base Price": 1012.3, "NSW / ACT": 1098.1, "NT": 1098.1, "QLD": 1012.3, "SA": 1182,   "TAS": 1182,   "VIC": 1012.3, "WA": 1012.3},
    {"Internal ID": 8242,  "Name": "ATF750.GASKET",        "Type": "Inventory Item",  "Display Name": "DN750 GASKET",                                  "Base Price": 85,     "NSW / ACT": 85,     "NT": 85,     "QLD": 85,     "SA": 85,     "TAS": 85,     "VIC": 85,     "WA": 85},
    {"Internal ID": 8154,  "Name": "ATF900.8",             "Type": "Assembly",        "Display Name": "AtlanFlow DN900 SN8",                           "Base Price": 1201.9, "NSW / ACT": 1629.8, "NT": 1629.8, "QLD": 1201.9, "SA": 1493,   "TAS": 1493,   "VIC": 1201.9, "WA": 1201.9},
    {"Internal ID": 8243,  "Name": "ATF900.GASKET",        "Type": "Inventory Item",  "Display Name": "DN900 GASKET",                                  "Base Price": 95,     "NSW / ACT": 95,     "NT": 95,     "QLD": 95,     "SA": 95,     "TAS": 95,     "VIC": 95,     "WA": 95},
    {"Internal ID": 8331,  "Name": "ATFJ100-225/300",      "Type": "Inventory Item",  "Display Name": "DN100 Joiner to 225/300 SN8 Atlan Flow",        "Base Price": 45,     "NSW / ACT": 45,     "NT": 45,     "QLD": 45,     "SA": 45,     "TAS": 45,     "VIC": 45,     "WA": 45},
    {"Internal ID": 8333,  "Name": "ATFJ100-375/450/525",  "Type": "Inventory Item",  "Display Name": "DN100 Joiner to 375/450/525 SN8 Atlan Flow",    "Base Price": 52,     "NSW / ACT": 52,     "NT": 52,     "QLD": 52,     "SA": 52,     "TAS": 52,     "VIC": 52,     "WA": 52},
    {"Internal ID": 8334,  "Name": "ATFJ150-300/375",      "Type": "Inventory Item",  "Display Name": "DN150 Joiner to 300/375 SN8 Atlan Flow",        "Base Price": 95,     "NSW / ACT": 95,     "NT": 95,     "QLD": 95,     "SA": 95,     "TAS": 95,     "VIC": 95,     "WA": 95},
    {"Internal ID": 8336,  "Name": "ATFJ150-375/450/525",  "Type": "Inventory Item",  "Display Name": "DN150 Joiner to 375/450/525 SN8 Atlan Flow",    "Base Price": 132,    "NSW / ACT": 132,    "NT": 132,    "QLD": 132,    "SA": 132,    "TAS": 132,    "VIC": 132,    "WA": 132},
    {"Internal ID": 8335,  "Name": "ATFJ150-600/750/900",  "Type": "Inventory Item",  "Display Name": "DN150 Joiner to 600/750/900 SN8 Atlan Flow",    "Base Price": 145,    "NSW / ACT": 145,    "NT": 145,     "QLD": 145,    "SA": 145,    "TAS": 145,    "VIC": 145,    "WA": 145},
]
BUILTIN_NETSUITE_DF = pd.DataFrame(_BUILTIN_PRICE_DATA)

# ---------------------------------------------------------------------------
# Competitor Intelligence expected schema (from the Power Apps / Power
# Automate build). Every row below is one pipe entry from one submission.
# ---------------------------------------------------------------------------
COMPETITOR_EXPECTED_COLUMNS = [
    "SubmittedBy", "QuoteDate", "ProjectName", "State", "Location", "Competitor",
    "AtlanReference", "LengthM", "Quantity", "TotalPrice", "Freight",
    "DeliveryLocation", "DispatchLocation",
]

# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------
st.markdown(
    """
<style>
.stApp { background: #F4F7FB; }
.block-container { max-width: 1450px; padding-top: 0.8rem; padding-bottom: 1rem; }
.hero {
    background: linear-gradient(135deg, #071B3A 0%, #0B5CFF 100%);
    padding: 22px 26px; border-radius: 22px; color: white;
    box-shadow: 0 14px 34px rgba(7,27,58,0.18); margin-bottom: 16px;
}
.hero h1 { font-size: 28px; margin-bottom: 4px; font-weight: 850; line-height: 1.1; }
.hero p { font-size: 13px; opacity: 0.90; margin-bottom: 0; line-height: 1.4; }
.card {
    background: white; border: 1px solid rgba(7,27,58,0.06); border-radius: 18px;
    padding: 16px; margin-bottom: 14px; box-shadow: 0 6px 18px rgba(7,27,58,0.04);
}
.title { font-size: 17px; font-weight: 800; color: #071B3A; margin-bottom: 2px; }
.subtle { color: rgba(7,27,58,0.60); font-size: 11px; line-height: 1.3; }
.good { background: #ECFDF3; }
.watch { background: #FFF7E6; }
.bad { background: #FFF1F1; }
.risk-box {
    border-radius: 16px; padding: 14px 16px; margin-top: 12px;
    border: 1px solid rgba(7,27,58,0.08); font-size: 12px;
}
[data-testid="stMetric"] {
    background: white; border-radius: 14px; padding: 10px 12px;
    border: 1px solid rgba(7,27,58,0.06); box-shadow: 0 4px 12px rgba(7,27,58,0.03);
}
[data-testid="stMetricLabel"] { font-size: 11px; color: rgba(7,27,58,0.62); }
[data-testid="stMetricValue"] { font-size: 21px; font-weight: 850; color: #071B3A; }
[data-testid="stMetricDelta"] { font-size: 11px; }
.stSelectbox label, .stNumberInput label, .stRadio label {
    font-size: 11px !important; font-weight: 650 !important;
}
.stTextInput input, .stNumberInput input { font-size: 12px !important; }
.stSelectbox div[data-baseweb="select"] { min-height: 34px; }
div.stButton > button { border-radius: 10px; font-weight: 700; font-size: 12px; padding: 0.35rem 0.8rem; }
div.stButton > button[kind="primary"] { background: #0B5CFF; border-color: #0B5CFF; }
section[data-testid="stSidebar"] { background: #FFFFFF; border-right: 1px solid rgba(7,27,58,0.08); }
[data-testid="stDataFrame"] { font-size: 11px; }
.streamlit-expanderHeader { font-size: 13px !important; font-weight: 700 !important; }
</style>
""",
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def money(value: float) -> str:
    return f"${value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def safe_divide(a: float, b: float) -> float:
    return a / b if b else 0.0


def format_display_df(df: pd.DataFrame, formats: Dict[str, str]) -> pd.DataFrame:
    """Pre-format numeric columns into plain display strings, instead of
    using pandas Styler. Streamlit renders Styler objects by serializing
    through Arrow, which has been an unstable path across some pandas/
    numpy/pyarrow version combinations (can segfault) — plain strings are
    always safe to render."""
    out = df.copy()
    for col, fmt in formats.items():
        if col not in out.columns:
            continue
        def _fmt(v, fmt=fmt):
            if pd.isna(v):
                return ""
            try:
                return fmt.format(v)
            except (ValueError, TypeError):
                return str(v)
        out[col] = out[col].apply(_fmt)
    return out


# ---------------------------------------------------------------------------
# Data loaders
# ---------------------------------------------------------------------------
def _read_spreadsheet(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Read xlsx, xls (including XML-disguised), or csv robustly."""
    name = filename.lower()
    if name.endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes))
    sniff = file_bytes[:200].lstrip()
    is_xml = sniff.startswith(b"<?xml") or sniff.startswith(b"<html") or b"<Workbook" in sniff
    if is_xml:
        tables = pd.read_html(io.BytesIO(file_bytes), header=0)
        return tables[0]
    if name.endswith(".xls"):
        return pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
    return pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")


@st.cache_data(show_spinner=False)
def load_netsuite(file_bytes: bytes, filename: str) -> pd.DataFrame:
    df = _read_spreadsheet(file_bytes, filename)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False)
def load_competitor(file_bytes: bytes, filename: str) -> pd.DataFrame:
    df = _read_spreadsheet(file_bytes, filename)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False, ttl=0)
def load_competitor_from_sharepoint() -> "tuple[Optional[pd.DataFrame], str]":
    """Auto-download Competitor Intelligence from SharePoint. Fetches fresh on every page load."""
    try:
        resp = requests.get(COMPETITOR_SHAREPOINT_URL, timeout=15)
        if resp.status_code == 200 and len(resp.content) > 1000:
            file_bytes = resp.content
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
            except Exception:
                try:
                    df = pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
                except Exception:
                    tables = pd.read_html(io.BytesIO(file_bytes), header=0)
                    df = tables[0]
            df.columns = [str(c).strip() for c in df.columns]
            return df, f"✓ Competitor data auto-loaded from SharePoint — {len(df):,} records"
        elif resp.status_code in (401, 403):
            return None, "SharePoint file requires Atlan login — upload manually below."
        else:
            return None, f"SharePoint returned status {resp.status_code} — upload manually below."
    except requests.exceptions.Timeout:
        return None, "SharePoint request timed out — upload manually below."
    except Exception as e:
        return None, f"Could not reach SharePoint ({e!s}) — upload manually below."


def get_item_price(netsuite_df: pd.DataFrame, item_name: str, region_key: str) -> Optional[float]:
    col = STATE_TO_NETSUITE_COL.get(region_key)
    if col is None or col not in netsuite_df.columns:
        col = "Base Price"
    if col not in netsuite_df.columns:
        return None
    mask = (
        netsuite_df.get("Display Name", pd.Series(dtype=str)).str.strip().str.lower() == item_name.strip().lower()
    ) | (
        netsuite_df.get("Name", pd.Series(dtype=str)).str.strip().str.lower() == item_name.strip().lower()
    )
    rows = netsuite_df[mask]
    if rows.empty:
        return None
    val = rows.iloc[0][col]
    try:
        v = float(val)
        if pd.isna(v):
            raise ValueError
        return v
    except (ValueError, TypeError):
        try:
            base = rows.iloc[0].get("Base Price", None)
            v = float(base)
            return None if pd.isna(v) else v
        except (ValueError, TypeError):
            return None


def get_item_code(netsuite_df: pd.DataFrame, display_name: str) -> Optional[str]:
    """Map a Display Name (shown in the item selector) back to its ATF code
    (the 'Name' column) — this is what competitor AtlanReference values are
    matched against, since the Power Apps form's dropdown uses these exact
    codes."""
    mask = netsuite_df.get("Display Name", pd.Series(dtype=str)).str.strip().str.lower() == display_name.strip().lower()
    rows = netsuite_df[mask]
    if rows.empty:
        return None
    code = rows.iloc[0].get("Name")
    return str(code).strip() if pd.notna(code) else None


def normalize_competitor_df(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce competitor data into the expected schema/types, tolerating
    minor naming drift (e.g. spaces) and missing optional columns."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    rename_map = {
        "Submitted By": "SubmittedBy",
        "Quote Date": "QuoteDate",
        "Project Name": "ProjectName",
        "Atlan Reference": "AtlanReference",
        "Competitor Pipe Size - Length (metres)": "LengthM",
        "Delivery Location": "DeliveryLocation",
        "Dispatch Location": "DispatchLocation",
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
    for col in COMPETITOR_EXPECTED_COLUMNS:
        if col not in df.columns:
            df[col] = None
    for col in ("LengthM", "Quantity", "TotalPrice", "Freight"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["State"] = df["State"].fillna("").astype(str).str.strip()
    df["AtlanReference"] = df["AtlanReference"].fillna("").astype(str).str.strip()
    df["Competitor"] = df["Competitor"].fillna("").astype(str).str.strip()
    # Price per metre for one pipe entry: TotalPrice covers Quantity pipes,
    # each LengthM metres long, so unit price/m = TotalPrice / (Qty * LengthM)
    denom = (df["Quantity"].fillna(0) * df["LengthM"].fillna(0)).replace(0, pd.NA)
    df["PricePerM"] = df["TotalPrice"] / denom
    # Price per UNIT (per pipe/each) — this is the primary metric used for
    # peer comparison, since Quantity is a count of physical units and
    # NetSuite prices are also per-unit, not per-metre. Comparing $/unit
    # avoids introducing error from length assumptions that may differ
    # between Atlan's stock length and what a competitor actually quoted.
    qty_denom = df["Quantity"].replace(0, pd.NA)
    df["PriceEach"] = df["TotalPrice"] / qty_denom
    return df


def get_competitor_rows_for_region(competitor_df: pd.DataFrame, region_key: str) -> pd.DataFrame:
    state_vals = COMPETITOR_STATE_MAP.get(region_key, [region_key])
    return competitor_df[competitor_df["State"].isin(state_vals)].copy()


def _extract_size_mm(code: str) -> Optional[str]:
    """Pull the numeric pipe size out of an ATF code, e.g. 'ATF225-45' and
    'ATF225.8' both yield '225'. Used only as a fallback when an exact code
    match isn't available."""
    m = re.search(r"ATF(\d{2,4})", str(code).upper())
    return m.group(1) if m else None


def get_competitor_price_for_code(comp_rows: pd.DataFrame, competitor_name: str, atlan_code: str) -> tuple[Optional[float], int, str]:
    """Match a competitor's submitted price-per-unit for a specific ATF
    code. Three tiers, tried in order:
    1. Exact reference match (e.g. 'ATF225.8' == 'ATF225.8')
    2. Same-size match (e.g. 'ATF225-45' fitting matches 'ATF225.8' pipe,
       both size 225) — approximate, comparing a fitting to a straight pipe
    3. Nearest-size match — if this competitor has no 225 at all, use
       whichever size they DO have that's numerically closest (e.g. 300)
       so there's always something usable rather than a blank comparison
    Returns (average $/unit, number of records matched, match_type) where
    match_type is 'exact', 'size_fallback', or 'nearest_size', or (None, 0,
    'none') if this competitor has no usable data at all."""
    exact = comp_rows[
        (comp_rows["Competitor"] == competitor_name)
        & (comp_rows["AtlanReference"].str.upper() == str(atlan_code).strip().upper())
    ]
    exact = exact.dropna(subset=["PriceEach"])
    if not exact.empty:
        return float(exact["PriceEach"].mean()), len(exact), "exact"

    target_size = _extract_size_mm(atlan_code)
    comp_subset = comp_rows[comp_rows["Competitor"] == competitor_name].copy()
    if comp_subset.empty:
        return None, 0, "none"
    comp_subset["_size"] = comp_subset["AtlanReference"].apply(_extract_size_mm)

    if target_size is not None:
        size_matches = comp_subset[comp_subset["_size"] == target_size].dropna(subset=["PriceEach"])
        if not size_matches.empty:
            return float(size_matches["PriceEach"].mean()), len(size_matches), "size_fallback"

    # Nearest-size fallback: this competitor has data, just not for this
    # exact size — use whichever size they do have that's numerically closest
    usable = comp_subset.dropna(subset=["PriceEach"]).copy()
    usable = usable[usable["_size"].notna()]
    if usable.empty or target_size is None:
        return None, 0, "none"
    usable["_size_num"] = usable["_size"].astype(float)
    target_num = float(target_size)
    usable["_distance"] = (usable["_size_num"] - target_num).abs()
    nearest_size = usable.loc[usable["_distance"].idxmin(), "_size"]
    nearest_matches = usable[usable["_size"] == nearest_size]
    return float(nearest_matches["PriceEach"].mean()), len(nearest_matches), "nearest_size"


# ---------------------------------------------------------------------------
# Freight
# ---------------------------------------------------------------------------
def calculate_freight(
    fleet_name: str,
    km_one_way: float,
    driver_rate: float,
    diesel_price: float,
    avg_speed: float,
    site_hours: float,
    trip_type: str,
    region_key: str,
) -> float:
    fleet = FLEET[fleet_name]
    region = REGIONS[region_key]
    total_km = km_one_way if trip_type == "One-way" else km_one_way * 2
    fuel_per_km = fleet.litres_per_100km / 100 * diesel_price
    vehicle_cost_per_km = fuel_per_km + fleet.maintenance_per_km
    drive_hours = safe_divide(total_km, avg_speed)
    labour_cost = (drive_hours + site_hours) * driver_rate
    vehicle_cost = total_km * vehicle_cost_per_km
    return (labour_cost + vehicle_cost) * region.freight_factor


# ---------------------------------------------------------------------------
# Session state helpers
# ---------------------------------------------------------------------------
def add_delivery() -> None:
    new_id = st.session_state.next_delivery_id
    st.session_state.next_delivery_id += 1
    st.session_state.deliveries.append(
        {
            "id": new_id,
            "products": [{
                "item_name": "", "rrp_per_m": 0.0, "quantity_m": 100.0, "discount_pct": 0,
                "stock_length_m": 6.0,
                "manual_override": False, "manual_rrp": 0.0, "manual_cost": 0.0,
            }],
            "freight_method": "Auto calculate",
            "zone": "Metro",
            "km_one_way": 30.0,
            "trip_type": "Return",
            "fleet": "6.5m truck",
            "site_hours": 1.0,
            "manual_freight": 0.0,
        }
    )


def remove_delivery(delivery_id: int) -> None:
    st.session_state.deliveries = [d for d in st.session_state.deliveries if d["id"] != delivery_id]


def add_product_to_delivery(delivery_id: int) -> None:
    for delivery in st.session_state.deliveries:
        if delivery["id"] == delivery_id:
            delivery["products"].append({
                "item_name": "", "rrp_per_m": 0.0, "quantity_m": 100.0, "discount_pct": 0,
                "stock_length_m": 6.0,
                "manual_override": False, "manual_rrp": 0.0, "manual_cost": 0.0,
            })
            break


def remove_product_from_delivery(delivery_id: int, product_index: int) -> None:
    for delivery in st.session_state.deliveries:
        if delivery["id"] == delivery_id and len(delivery["products"]) > 1:
            delivery["products"].pop(product_index)
            break


# ---------------------------------------------------------------------------
# Calculation
# ---------------------------------------------------------------------------
def calculate_delivery(delivery: dict, global_inputs: dict, region_key: str) -> tuple[list[dict], float]:
    if delivery["freight_method"] == "Manual override":
        delivery_freight = delivery["manual_freight"]
    else:
        delivery_freight = calculate_freight(
            fleet_name=delivery["fleet"],
            km_one_way=delivery["km_one_way"],
            driver_rate=global_inputs["driver_rate"],
            diesel_price=global_inputs["diesel_price"],
            avg_speed=global_inputs["avg_speed"],
            site_hours=delivery["site_hours"],
            trip_type=delivery["trip_type"],
            region_key=region_key,
        )

    temp_rows = []
    total_delivery_revenue = 0.0
    for product in delivery["products"]:
        item_name = product.get("item_name", "")
        quantity_m = product["quantity_m"]
        discount_pct = product["discount_pct"]

        if product.get("manual_override"):
            rrp_per_m = product.get("manual_rrp", 0.0) or 0.0
            cost_per_m = product.get("manual_cost", 0.0) or 0.0
        else:
            stock_length_m = product.get("stock_length_m", 6.0) or 6.0
            raw_each_price = resolve_price(item_name) if item_name else (product.get("rrp_per_m", 0.0) or 0.0)
            rrp_per_m = safe_divide(raw_each_price, stock_length_m)
            cost_per_m = round(rrp_per_m * COST_FACTOR, 4)

        # Stock length is needed regardless of override, to convert this
        # line's $/m figures into $/unit for peer comparison (competitors
        # are compared by unit price, not $/m — see build_peer_comparison)
        stock_length_m_for_units = product.get("stock_length_m", 6.0) or 6.0

        net_price_per_m = rrp_per_m * (1 - discount_pct / 100)
        rrp_revenue = rrp_per_m * quantity_m
        revenue = net_price_per_m * quantity_m
        product_cost = cost_per_m * quantity_m
        total_delivery_revenue += revenue

        units = safe_divide(quantity_m, stock_length_m_for_units)

        temp_rows.append(
            {
                "Delivery": f"Delivery {delivery['id']}",
                "Item": item_name,
                "Quantity m": quantity_m,
                "Stock Length (m)": stock_length_m_for_units,
                "Units": units,
                "RRP / m": rrp_per_m,
                "Cost / m": cost_per_m,
                "Discount %": discount_pct,
                "Net Price / m": net_price_per_m,
                "Net Price Each": net_price_per_m * stock_length_m_for_units,
                "RRP Each": rrp_per_m * stock_length_m_for_units,
                "RRP Revenue": rrp_revenue,
                "Revenue": revenue,
                "Product Cost": product_cost,
                "Manual Override": bool(product.get("manual_override")),
            }
        )

    final_rows = []
    for row in temp_rows:
        allocation_pct = safe_divide(row["Revenue"], total_delivery_revenue)
        freight_allocated = delivery_freight * allocation_pct
        total_cost = row["Product Cost"] + freight_allocated
        contribution = row["Revenue"] - total_cost
        contribution_margin = safe_divide(contribution, row["Revenue"])
        rrp_contribution = row["RRP Revenue"] - total_cost
        rrp_margin = safe_divide(rrp_contribution, row["RRP Revenue"])
        margin_lost = rrp_contribution - contribution
        margin_lost_pp = (rrp_margin - contribution_margin) * 100
        row.update(
            {
                "Freight Allocated": freight_allocated,
                "Total Cost": total_cost,
                "Contribution $": contribution,
                "Contribution Margin %": contribution_margin,
                "RRP Contribution $": rrp_contribution,
                "RRP Margin %": rrp_margin,
                "Margin Lost $": margin_lost,
                "Margin Lost pp": margin_lost_pp,
            }
        )
        final_rows.append(row)
    return final_rows, delivery_freight


def build_peer_comparison(
    detail_df: pd.DataFrame,
    netsuite_df: pd.DataFrame,
    competitor_rows: pd.DataFrame,
    total_revenue: float,
    total_freight: float,
    peer_freight: Dict[str, float] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compares Atlan's package against each competitor on a PER-UNIT basis
    (price per pipe), not per-metre. Quantity is a count of physical units
    and NetSuite/competitor prices are both quoted per unit, so comparing
    $/unit avoids introducing error from stock-length assumptions that may
    not match what a competitor actually quoted."""
    if peer_freight is None:
        peer_freight = {}

    detail_df = detail_df.copy()
    detail_df["_atlan_code"] = detail_df["Item"].apply(lambda n: get_item_code(netsuite_df, n) or "")

    atlan_units = detail_df["Units"].sum()
    atlan_package = sum(line["Net Price Each"] * line["Units"] for _, line in detail_df.iterrows())

    competitor_names = sorted(competitor_rows["Competitor"].dropna().unique().tolist())
    summary_rows = []
    line_rows = []

    for comp_name in competitor_names:
        comp_package = 0.0
        matched_any = False
        for _, line in detail_df.iterrows():
            atlan_code = line["_atlan_code"]
            units = line["Units"]
            atlan_net_each = line["Net Price Each"]
            atlan_line_total = atlan_net_each * units

            comp_price_each, n_records, match_type = get_competitor_price_for_code(competitor_rows, comp_name, atlan_code)
            if comp_price_each is None:
                comp_price_each = 0.0
            else:
                matched_any = True
            comp_line_total = comp_price_each * units
            comp_package += comp_line_total

            line_rows.append({
                "Competitor": comp_name,
                "Item": line["Item"],
                "Atlan Reference": atlan_code,
                "Comp. $/Unit": comp_price_each,
                "Match Type": {
                    "exact": "Exact", "size_fallback": "Approx (same size)",
                    "nearest_size": "Approx (nearest size)", "none": "No match",
                }[match_type],
                "Records Matched": n_records,
                "Atlan Net $/Unit": atlan_net_each,
                "Units": units,
                "Comp. Line Total": comp_line_total,
                "Atlan Line Total": atlan_line_total,
                "Line $ Diff": atlan_line_total - comp_line_total,
            })

        if not matched_any:
            continue  # this competitor has no data for any item in the package

        comp_freight = peer_freight.get(comp_name, total_freight)
        summary_rows.append({
            "Supplier": comp_name,
            "Product Package": comp_package,
            "Freight": comp_freight,
            "Total Package": comp_package + comp_freight,
            "Avg $/Unit": safe_divide(comp_package, atlan_units),
        })

    summary_rows.append({
        "Supplier": "✦ Atlan Proposed",
        "Product Package": atlan_package,
        "Freight": total_freight,
        "Total Package": atlan_package + total_freight,
        "Avg $/Unit": safe_divide(atlan_package, atlan_units),
    })

    summary_df = pd.DataFrame(summary_rows).sort_values("Total Package").reset_index(drop=True)
    line_df = pd.DataFrame(line_rows)
    return summary_df, line_df


# ---------------------------------------------------------------------------
# Excel export (formatted, multi-tab)
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="0B5CFF", end_color="0B5CFF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="071B3A", bold=True, size=14)


def _style_header_row(ws, n_cols: int, row: int = 1) -> None:
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_columns(ws, df: pd.DataFrame, start_col: int = 1) -> None:
    for i, col in enumerate(df.columns):
        col_letter = get_column_letter(start_col + i)
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)


def _write_df_sheet(writer, df: pd.DataFrame, sheet_name: str, currency_cols: list[str] = None, title: str = None) -> None:
    currency_cols = currency_cols or []
    start_row = 2 if title else 0
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)
    ws = writer.sheets[sheet_name]
    if title:
        ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    _style_header_row(ws, len(df.columns), row=start_row + 1)
    _autosize_columns(ws, df)
    for col_name in currency_cols:
        if col_name not in df.columns:
            continue
        col_idx = df.columns.get_loc(col_name) + 1
        for r in range(start_row + 2, start_row + 2 + len(df)):
            ws.cell(row=r, column=col_idx).number_format = '$#,##0.00'
    ws.freeze_panes = ws.cell(row=start_row + 2, column=1)


def build_excel_export(
    detail_df: pd.DataFrame,
    competitor_rows_region: pd.DataFrame,
    summary_df: pd.DataFrame,
    line_df: pd.DataFrame,
    region_key: str,
) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_df_sheet(
            writer, detail_df, "Package Detail",
            currency_cols=["RRP / m", "Cost / m", "Net Price / m", "Net Price Each", "RRP Each",
                           "RRP Revenue", "Revenue", "Product Cost", "Freight Allocated",
                           "Total Cost", "Contribution $", "RRP Contribution $", "Margin Lost $"],
            title=f"Atlan Package Detail — {region_key}",
        )
        _write_df_sheet(
            writer, summary_df, "Peer Comparison Summary",
            currency_cols=["Product Package", "Freight", "Total Package", "Avg $/Unit"],
            title=f"Atlan vs Competitors — Total Package Comparison ({region_key})",
        )
        if not line_df.empty:
            _write_df_sheet(
                writer, line_df, "Line-by-Line Breakdown",
                currency_cols=["Comp. $/Unit", "Atlan Net $/Unit", "Comp. Line Total", "Atlan Line Total", "Line $ Diff"],
                title="Line-by-Line: Atlan vs Each Competitor, per Item",
            )
        display_cols = [c for c in COMPETITOR_EXPECTED_COLUMNS if c in competitor_rows_region.columns] + (
            ["PricePerM"] if "PricePerM" in competitor_rows_region.columns else []
        )
        _write_df_sheet(
            writer, competitor_rows_region[display_cols], "Competitor Raw Data",
            currency_cols=["TotalPrice", "Freight", "PriceEach", "PricePerM"],
            title=f"Raw Competitor Submissions — {region_key}",
        )
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Session state init
# ---------------------------------------------------------------------------
if "deliveries" not in st.session_state:
    st.session_state.deliveries = []
if "next_delivery_id" not in st.session_state:
    st.session_state.next_delivery_id = 1
if not st.session_state.deliveries:
    add_delivery()

# ---------------------------------------------------------------------------
# Hero
# ---------------------------------------------------------------------------
st.markdown(
    """
<div class="hero">
    <h1>Atlan Stormwater Pricing Engine</h1>
    <p>
        Build a multi-delivery pipe package, apply controlled discounts, calculate freight by delivery,
        and compare Atlan's total package against competitor intelligence.
    </p>
</div>
""",
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Sidebar
# ---------------------------------------------------------------------------
with st.sidebar:
    st.markdown("### Data Sources")
    netsuite_file = st.file_uploader(
        "NetSuite Price List — optional override (.xlsx / .csv)",
        type=["xlsx", "xls", "csv"],
        key="netsuite_upload",
        help="Upload a fresh NetSuite export to override the built-in ATF price list. Same column format required.",
    )
    if netsuite_file:
        st.session_state["_netsuite_bytes"] = (netsuite_file.read(), netsuite_file.name)

    netsuite_df: Optional[pd.DataFrame] = None
    competitor_df: Optional[pd.DataFrame] = None
    netsuite_source = "built-in ATF price list"

    if "_netsuite_bytes" in st.session_state:
        try:
            b, name = st.session_state["_netsuite_bytes"]
            netsuite_df = load_netsuite(b, name)
            netsuite_source = f"uploaded file ({name})"
            st.success(f"✓ NetSuite override loaded — {len(netsuite_df):,} items")
        except Exception as e:
            st.error(f"Failed to load NetSuite file: {e}")
            netsuite_df = BUILTIN_NETSUITE_DF
    else:
        netsuite_df = BUILTIN_NETSUITE_DF
        st.info(f"✓ Using built-in ATF price list ({len(netsuite_df):,} items). Upload a file above to override.")

    st.markdown("**Competitor Intelligence**")
    sp_df, sp_msg = load_competitor_from_sharepoint()
    if sp_df is not None:
        competitor_df = normalize_competitor_df(sp_df)
        st.success(sp_msg)
        st.caption("Fetches the latest approved submissions from SharePoint on every page load. Upload below to override.")
    else:
        st.warning(sp_msg)

    competitor_file = st.file_uploader(
        "Manual override — Competitor Intelligence (.xlsx / .csv)",
        type=["xlsx", "xls", "csv"],
        key="competitor_upload",
        help="Only needed if SharePoint auto-load fails. Expected columns: SubmittedBy, QuoteDate, ProjectName, "
             "State, Location, Competitor, AtlanReference, LengthM, Quantity, TotalPrice, Freight, "
             "DeliveryLocation, DispatchLocation.",
    )
    if competitor_file:
        try:
            b = competitor_file.read()
            raw_df = load_competitor(b, competitor_file.name)
            competitor_df = normalize_competitor_df(raw_df)
            st.success(f"✓ Manual upload loaded — {len(competitor_df):,} records (overrides SharePoint)")
        except Exception as e:
            st.error(f"Failed to load competitor file: {e}")

    st.divider()
    st.markdown("### Market & Region")
    region_key = st.selectbox(
        "State / Region",
        list(REGIONS.keys()),
        format_func=lambda x: f"{x} — {REGIONS[x].name}",
    )
    st.caption(REGIONS[region_key].notes)

    st.divider()
    st.markdown("### Freight Inputs")
    driver_rate = st.number_input("Driver $ / hr", min_value=0.0, value=100.0, step=5.0)
    diesel_price = st.number_input("Diesel $ / L", min_value=0.0, value=3.00, step=0.10)
    avg_speed = st.number_input("Average km / h", min_value=1.0, value=60.0, step=5.0)

    st.divider()
    st.markdown("### Guardrails")
    target_margin = st.slider("Target margin %", 0, 70, 35, 1) / 100
    risk_margin = st.slider("High-risk margin %", 0, 50, 25, 1) / 100

global_inputs = {"driver_rate": driver_rate, "diesel_price": diesel_price, "avg_speed": avg_speed}

# ---------------------------------------------------------------------------
# Build item options for selectors
# ---------------------------------------------------------------------------
price_col = STATE_TO_NETSUITE_COL.get(region_key, "Base Price")
if price_col not in netsuite_df.columns:
    price_col = "Base Price"
name_col = "Display Name" if "Display Name" in netsuite_df.columns else "Name"
price_series = pd.to_numeric(netsuite_df.get(price_col, netsuite_df.get("Base Price", pd.Series(dtype=float))), errors="coerce")
base_series = pd.to_numeric(netsuite_df.get("Base Price", pd.Series(dtype=float)), errors="coerce")
effective_price = price_series.combine_first(base_series)
item_mask = effective_price.notna() & (effective_price > 0)
item_names = netsuite_df.loc[item_mask, name_col].dropna().str.strip().sort_values().tolist()


def resolve_price(item_name: str) -> float:
    """Return the raw NetSuite sell price for item_name in current region.
    This is priced per EACH pipe (per stock length, e.g. ~6m) — NOT per
    metre. Callers must divide by stock length to get $/m; see
    calculate_delivery, which does this via product['stock_length_m']."""
    p = get_item_price(netsuite_df, item_name, region_key)
    return p if p is not None else 0.0


# ---------------------------------------------------------------------------
# Package Builder
# ---------------------------------------------------------------------------
top_left, top_right = st.columns([0.78, 0.22])
with top_left:
    st.markdown("### Package Builder")
    st.caption(
        f"Prices from {netsuite_source} — {STATE_TO_NETSUITE_COL.get(region_key, 'Base Price')} column. "
        "Each delivery can include multiple items. Freight is allocated across products. "
        "Tick 'Manual override' on any line to type your own price/cost instead of the list price."
    )
with top_right:
    if st.button("+ Add Delivery", type="primary", width="stretch"):
        add_delivery()
        st.rerun()

all_rows = []
for delivery in list(st.session_state.deliveries):
    st.markdown('<div class="card">', unsafe_allow_html=True)
    h1, h2 = st.columns([0.82, 0.18])
    with h1:
        st.markdown(f'<div class="title">Delivery {delivery["id"]}</div>', unsafe_allow_html=True)
        st.markdown('<div class="subtle">Select items from the price list, enter quantity and discount.</div>', unsafe_allow_html=True)
    with h2:
        if len(st.session_state.deliveries) > 1:
            if st.button("Remove", key=f"remove_delivery_{delivery['id']}", width="stretch"):
                remove_delivery(delivery["id"])
                st.rerun()

    for idx, product in enumerate(list(delivery["products"])):
        p1, p2, p3, p4, p5, p6, p7 = st.columns([0.20, 0.10, 0.11, 0.13, 0.13, 0.12, 0.14])
        with p1:
            current_item = product.get("item_name", "")
            if current_item not in item_names:
                current_item = item_names[0] if item_names else ""
            try:
                default_idx = item_names.index(current_item)
            except ValueError:
                default_idx = 0
            selected_item = st.selectbox(
                "Item",
                item_names,
                index=default_idx,
                key=f"item_{delivery['id']}_{idx}",
            )
            product["item_name"] = selected_item

        with p2:
            product["quantity_m"] = st.number_input(
                "Qty m",
                min_value=0.0,
                value=float(product["quantity_m"]),
                step=10.0,
                key=f"qty_{delivery['id']}_{idx}",
            )

        product["manual_override"] = st.checkbox(
            "Manual override price / cost",
            value=product.get("manual_override", False),
            key=f"override_{delivery['id']}_{idx}",
        )

        raw_each_price = resolve_price(product["item_name"]) if product.get("item_name") else 0.0

        if product["manual_override"]:
            with p3:
                st.metric("List (each)", f"${raw_each_price:,.2f}")
            with p4:
                product["manual_rrp"] = st.number_input(
                    "Price/m ($)",
                    min_value=0.0,
                    value=float(product.get("manual_rrp") or 0.0),
                    step=1.0,
                    key=f"manual_rrp_{delivery['id']}_{idx}",
                )
            with p5:
                product["manual_cost"] = st.number_input(
                    "Cost/m ($)",
                    min_value=0.0,
                    value=float(product.get("manual_cost") or (product["manual_rrp"] * COST_FACTOR)),
                    step=1.0,
                    key=f"manual_cost_{delivery['id']}_{idx}",
                )
            current_price = product["manual_rrp"]
            current_cost = product["manual_cost"]
        else:
            with p3:
                product["stock_length_m"] = st.number_input(
                    "Stock length (m)",
                    min_value=0.1,
                    value=float(product.get("stock_length_m", 6.0)),
                    step=0.1,
                    key=f"stock_len_{delivery['id']}_{idx}",
                    help="NetSuite lists price per pipe (each), not per metre. This converts it: "
                         "Price/m = list price ÷ stock length.",
                )
            stock_length_m = product["stock_length_m"]
            current_price = safe_divide(raw_each_price, stock_length_m)
            current_cost = current_price * COST_FACTOR
            product["rrp_per_m"] = current_price
            with p4:
                st.metric("Price/m", f"${current_price:,.2f}", delta=f"${raw_each_price:,.2f} each")
            with p5:
                st.metric("Cost/m", f"${current_cost:,.2f}")

        with p6:
            product["discount_pct"] = st.selectbox(
                "Discount",
                DISCOUNT_OPTIONS,
                index=DISCOUNT_OPTIONS.index(product["discount_pct"]),
                key=f"discount_{delivery['id']}_{idx}",
                format_func=lambda x: f"{x}%",
            )
        with p7:
            net = current_price * (1 - product["discount_pct"] / 100)
            margin_preview = safe_divide(net - current_cost, net)
            st.metric("Net/m", f"${net:,.2f}", delta=f"{margin_preview:.1%} margin")
            if len(delivery["products"]) > 1:
                if st.button("✕ Remove line", key=f"remove_product_{delivery['id']}_{idx}", width="stretch"):
                    remove_product_from_delivery(delivery["id"], idx)
                    st.rerun()

    if st.button("+ Add Item", key=f"add_product_{delivery['id']}", width="stretch"):
        add_product_to_delivery(delivery["id"])
        st.rerun()

    with st.expander("Freight Settings", expanded=False):
        f1, f2, f3, f4 = st.columns(4)
        with f1:
            delivery["freight_method"] = st.radio(
                "Method", ["Auto calculate", "Manual override"], horizontal=True, key=f"freight_method_{delivery['id']}"
            )
        with f2:
            delivery["trip_type"] = st.radio(
                "Trip", ["Return", "One-way"], horizontal=True, key=f"trip_{delivery['id']}"
            )
        with f3:
            delivery["zone"] = st.selectbox(
                "Zone", list(ZONES.keys()), index=list(ZONES.keys()).index(delivery["zone"]), key=f"zone_{delivery['id']}"
            )
        with f4:
            if st.button("Use Zone km", key=f"use_zone_{delivery['id']}", width="stretch"):
                delivery["km_one_way"] = float(ZONES[delivery["zone"]])
                st.rerun()
        f5, f6, f7, f8 = st.columns(4)
        with f5:
            delivery["km_one_way"] = st.number_input(
                "One-way km", min_value=0.0, value=float(delivery["km_one_way"]), step=10.0, key=f"km_{delivery['id']}"
            )
        with f6:
            delivery["fleet"] = st.selectbox(
                "Fleet", list(FLEET.keys()), index=list(FLEET.keys()).index(delivery["fleet"]), key=f"fleet_{delivery['id']}"
            )
        with f7:
            delivery["site_hours"] = st.number_input(
                "Site hrs", min_value=0.0, value=float(delivery["site_hours"]), step=0.5, key=f"site_hours_{delivery['id']}"
            )
        with f8:
            if delivery["freight_method"] == "Manual override":
                delivery["manual_freight"] = st.number_input(
                    "Manual freight $", min_value=0.0, value=float(delivery["manual_freight"]), step=50.0, key=f"manual_freight_{delivery['id']}"
                )

    delivery_rows, delivery_freight = calculate_delivery(delivery, global_inputs, region_key)
    all_rows.extend(delivery_rows)

    delivery_revenue = sum(r["Revenue"] for r in delivery_rows)
    delivery_contribution = sum(r["Contribution $"] for r in delivery_rows)
    delivery_margin = safe_divide(delivery_contribution, delivery_revenue)
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Revenue", money(delivery_revenue))
    m2.metric("Freight", money(delivery_freight))
    m3.metric("Contribution", money(delivery_contribution))
    m4.metric("Margin", pct(delivery_margin))
    m5.metric("Lines", len(delivery["products"]))
    st.markdown("</div>", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
detail_df = pd.DataFrame(all_rows)
if detail_df.empty:
    st.warning("Please add at least one product line.")
    st.stop()

total_quantity = detail_df["Quantity m"].sum()
total_rrp_revenue = detail_df["RRP Revenue"].sum()
total_revenue = detail_df["Revenue"].sum()
total_product_cost = detail_df["Product Cost"].sum()
total_freight = detail_df["Freight Allocated"].sum()
total_contribution = detail_df["Contribution $"].sum()
package_margin = safe_divide(total_contribution, total_revenue)
rrp_contribution = detail_df["RRP Contribution $"].sum()
rrp_margin = safe_divide(rrp_contribution, total_rrp_revenue)
weighted_discount = safe_divide(total_rrp_revenue - total_revenue, total_rrp_revenue)
margin_lost = rrp_contribution - total_contribution
margin_lost_pp = (rrp_margin - package_margin) * 100

st.markdown("### Executive Summary")
st.markdown('<div class="card">', unsafe_allow_html=True)
s1, s2, s3, s4, s5 = st.columns(5)
s1.metric("Revenue", money(total_revenue), delta=f"{pct(weighted_discount)} discount")
s2.metric("Contribution", money(total_contribution))
s3.metric("Margin", pct(package_margin))
s4.metric("Margin at Risk", money(margin_lost), delta=f"{margin_lost_pp:.1f} pts")
s5.metric("Freight", money(total_freight))
s6, s7, s8, s9, s10 = st.columns(5)
s6.metric("List Revenue", money(total_rrp_revenue))
s7.metric("Product Cost", money(total_product_cost))
s8.metric("Quantity", f"{total_quantity:,.0f}m")
s9.metric("List Margin", pct(rrp_margin))
s10.metric("Lines", len(detail_df))

if package_margin < risk_margin:
    risk_class, risk_title = "bad", "High margin risk"
    risk_message = "Below the high-risk threshold. Review discounting, freight recovery or cost."
elif package_margin < target_margin:
    risk_class, risk_title = "watch", "Margin below target"
    risk_message = "Above the risk floor but below target. Check whether the discount is justified."
else:
    risk_class, risk_title = "good", "Healthy package margin"
    risk_message = "Above the target contribution margin."

st.markdown(
    f"""
<div class="risk-box {risk_class}">
    <b>{risk_title}</b><br>
    {risk_message}
    At list price, margin would be <b>{rrp_margin:.1%}</b>.
    After discount and freight, it is <b>{package_margin:.1%}</b>.
    Contribution at risk is <b>{money(margin_lost)}</b>.
</div>
""",
    unsafe_allow_html=True,
)
st.markdown("</div>", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Competitor comparison
# ---------------------------------------------------------------------------
st.markdown("### Competitor Intelligence")
st.markdown('<div class="card">', unsafe_allow_html=True)

summary_df = pd.DataFrame()
line_df = pd.DataFrame()
competitor_rows_region = pd.DataFrame()

try:
    if competitor_df is not None:
        competitor_rows_region = get_competitor_rows_for_region(competitor_df, region_key)
        comp_names = sorted(competitor_rows_region["Competitor"].dropna().unique().tolist())

        if comp_names:
            with st.expander("Peer Freight Assumptions", expanded=False):
                st.caption("Competitor freight defaults to Atlan's calculated freight. Edit each manually if known.")
                peer_freight = {}
                peer_cols = st.columns(len(comp_names))
                for col, comp_name in zip(peer_cols, comp_names):
                    with col:
                        peer_freight[comp_name] = st.number_input(
                            comp_name,
                            min_value=0.0,
                            value=float(total_freight),
                            step=50.0,
                            key=f"peer_freight_{comp_name}",
                        )

            summary_df, line_df = build_peer_comparison(
                detail_df, netsuite_df, competitor_rows_region, total_revenue, total_freight, peer_freight
            )

            st.caption(
                f"Competitor prices filtered to {region_key} / {', '.join(COMPETITOR_STATE_MAP.get(region_key, [region_key]))} region. "
                f"{len(competitor_rows_region):,} approved records available. "
                f"Compared on a **per-unit** basis (price per pipe), not per-metre — matching tries the exact "
                f"Atlan Reference code first (e.g. ATF300.8), then falls back to the same size (fitting vs pipe), "
                f"then to whichever size that competitor DOES have that's numerically closest, if needed."
            )

            if not summary_df.empty:
                st.dataframe(
                    format_display_df(summary_df, {
                        "Avg $/Unit": "${:,.2f}",
                        "Product Package": "${:,.0f}",
                        "Freight": "${:,.0f}",
                        "Total Package": "${:,.0f}",
                    }),
                    width="stretch",
                    hide_index=True,
                )
                comp_packages = summary_df.loc[summary_df["Supplier"] != "✦ Atlan Proposed", "Total Package"]
                if not comp_packages.empty:
                    peer_avg = comp_packages.mean()
                    gap = safe_divide(total_revenue - peer_avg, peer_avg)
                    if gap > 0.10:
                        st.warning(f"Atlan is priced {gap:.1%} above the competitor average package.")
                    elif gap < -0.05:
                        st.success(f"Atlan is priced {abs(gap):.1%} below the competitor average package.")
                    else:
                        st.info(f"Atlan is broadly market-aligned at {gap:.1%} versus the competitor average.")
            else:
                st.info("No competitor has approved pricing for any of the Atlan Reference codes in this package yet.")

            if not line_df.empty:
                with st.expander("Line-by-line competitor breakdown", expanded=False):
                    st.caption("For each Atlan line item, matched against competitor submissions on a per-unit basis — check 'Match Type' to see whether it's an exact code match, a same-size approximation, or a nearest-available-size approximation.")
                    fmt = {
                        "Comp. $/Unit": "${:,.2f}",
                        "Atlan Net $/Unit": "${:,.2f}",
                        "Units": "{:,.1f}",
                        "Comp. Line Total": "${:,.0f}",
                        "Atlan Line Total": "${:,.0f}",
                        "Line $ Diff": "${:,.0f}",
                    }
                    st.dataframe(format_display_df(line_df, fmt), width="stretch", hide_index=True)

            with st.expander("Raw competitor records for this region", expanded=False):
                display_cols = [c for c in COMPETITOR_EXPECTED_COLUMNS if c in competitor_rows_region.columns] + ["PriceEach", "PricePerM"]
                st.dataframe(
                    format_display_df(competitor_rows_region[display_cols], {"PriceEach": "${:,.2f}", "PricePerM": "${:,.2f}", "TotalPrice": "${:,.0f}", "Freight": "${:,.0f}"}),
                    width="stretch",
                    hide_index=True,
                )
        else:
            st.info(f"No competitor records found for region **{region_key}** in the loaded file.")
    else:
        st.info("Upload the Competitor Intelligence file in the sidebar to enable peer comparison.")
except Exception as e:
    st.error(f"Competitor comparison could not be built: {e}")
    summary_df = pd.DataFrame()
    line_df = pd.DataFrame()
    competitor_rows_region = pd.DataFrame()
st.markdown("</div>", unsafe_allow_html=True)

# ---------------------------------------------------------------------------
# Detailed output
# ---------------------------------------------------------------------------
with st.expander("Detailed Product Output", expanded=False):
    try:
        st.dataframe(
            format_display_df(detail_df, {
                "Quantity m": "{:,.0f}",
                "Stock Length (m)": "{:,.2f}",
                "Units": "{:,.1f}",
                "RRP / m": "${:,.2f}",
                "Cost / m": "${:,.2f}",
                "Discount %": "{:.0f}%",
                "Net Price / m": "${:,.2f}",
                "Net Price Each": "${:,.2f}",
                "RRP Each": "${:,.2f}",
                "RRP Revenue": "${:,.0f}",
                "Revenue": "${:,.0f}",
                "Product Cost": "${:,.0f}",
                "Freight Allocated": "${:,.0f}",
                "Total Cost": "${:,.0f}",
                "Contribution $": "${:,.0f}",
                "Contribution Margin %": "{:.1%}",
                "RRP Contribution $": "${:,.0f}",
                "RRP Margin %": "{:.1%}",
                "Margin Lost $": "${:,.0f}",
                "Margin Lost pp": "{:.1f} pts",
            }),
            width="stretch",
            hide_index=True,
        )
    except Exception as e:
        st.warning(f"Could not apply formatting to the detail table ({e}); showing raw data instead.")
        st.dataframe(detail_df, width="stretch", hide_index=True)

# ---------------------------------------------------------------------------
# Downloads
# ---------------------------------------------------------------------------
dl1, dl2 = st.columns(2)
with dl1:
    csv = detail_df.to_csv(index=False)
    st.download_button(
        label="Download package detail (CSV)",
        data=csv,
        file_name="atlan_pricing_output.csv",
        mime="text/csv",
        width="stretch",
    )
with dl2:
    try:
        excel_bytes = build_excel_export(detail_df, competitor_rows_region, summary_df, line_df, region_key)
        st.download_button(
            label="Download full comparison (Excel, formatted, multi-tab)",
            data=excel_bytes,
            file_name=f"atlan_competitor_comparison_{region_key}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
    except Exception as e:
        st.error(f"Could not build the Excel export: {e}")
